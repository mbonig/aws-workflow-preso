import json

# def lambda_handler(event, context):
#     print("Received event: " + json.dumps(event, indent=2))
#     # TODO implement
#     message = 'Hello {} {}!'.format(event['first_name'], 
#                                     event['last_name'])  
#     return {
#         'statusCode': 200,
#         'body': message
#     }
import boto3
import os
from io import BytesIO
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

from fpdf import FPDF

def lambda_handler(event, context):
    print("Received body: " + event['body'])
    doc = json.loads(event['body'])

    # TODO implement
    message = 'Hello {} {}!'.format(doc['first_name'], doc['last_name'])
    runExcel("test.xlsx")
    runPdf("test.pdf", doc)
    #print("output")
    return {
        'statusCode': 200,
        'body': message
    }

def runPdf(fileName, doc):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_xy(0, 0)
    pdf.set_font('arial', 'B', 13.0)
    pdf.cell(ln=0, h=5.0, align='L', w=0, txt=f"Hello {doc['first_name']} {doc['last_name']}!", border=0)
    pdfOut = pdf.output(name=fileName, dest='S').encode('latin-1')
    output = BytesIO(pdfOut)
    s3 = boto3.resource(
        's3'
    )
    bucket_name = os.environ['S3_BUCKET']
    print(f"bucket_name: {bucket_name}")
    s3.Bucket(bucket_name).upload_fileobj(output, fileName)
    #target_object = s3.Object(bucket_name, fileName)
    #target_object.put(Body=output)
def runExcel(fileName):
    wb = Workbook()
    #with open(fileName, 'wb') as file:
    ws1 = wb.active
    #print("file Name: " + file.name)
    wb = Workbook()
    ws1.title = "range names"
    for row in range(1, 40):
        ws1.append(range(600))
        
    ws2 = wb.create_sheet(title="Pi")
    ws2['F5'] = 3.14
    ws3 = wb.create_sheet(title="Data")
    for row in range(10, 20):
        for col in range(27, 54):
            _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
    print(ws3['AA10'].value)
    #output = BytesIO(file)
    #wb.save(file)
    output = BytesIO(save_virtual_workbook(wb))
    # with NamedTemporaryFile(delete=False) as tmp:
    #     wb.save(tmp.name)
    #     output = BytesIO(tmp.read())

    s3 = boto3.resource(
        's3'
    )
    bucket_name = os.environ['S3_BUCKET']

    s3.Bucket(bucket_name).upload_fileobj(output, fileName)
    #target_object = s3.Object(bucket_name,fileName)
    #target_object.put(Body=output)
        #fileobj = BytesIO(file.read())
        #s3.upload_fileobj(output, 'awsserverless-robk', fileName)